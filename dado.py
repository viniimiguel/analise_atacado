from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException
import openpyxl


class Atacadao():
    def __init__(self):
        self.driver = webdriver.Chrome()
        self.site_link = 'https://www.atacadao.com.br/mercearia/'
        self.cidade = '55660000'
        self.site_map = {
            'XP':{
                'cep': '/html/body/div[7]/div/div/div/div/div/div/div[2]/div/div[2]/div[2]/div/div[1]/div/div/input',
                'confirmar': '/html/body/div[7]/div/div/div/div/div/div/div[2]/div/div[3]/button[2]'
            }
        }

    def main(self):
        self.abre()
        sleep(1)
        self.cep()
        sleep(3)
        self.raspa()
        sleep(3)
        self.cria_planilhas()
        sleep(12312312)

    def abre(self):
        self.driver.get(self.site_link)
        self.driver.maximize_window()

    def cep(self):
        try:
            cidade = self.driver.find_element(By.XPATH, self.site_map['XP']['cep'])
            sleep(1)
            cidade.send_keys(self.cidade)
            sleep(3)
            self.driver.find_element(By.XPATH, self.site_map['XP']['confirmar']).click()
            sleep(5)
        except NoSuchElementException:
            print('elementto nao  encontrado tentando novamente...')
            self.cep()
        except Exception as e:
            print(f'error {e}')

    def desce(self):
        actions = ActionChains(self.driver)
        actions.send_keys(Keys.PAGE_DOWN).perform()

    def raspa(self):
        global armazena_nome,armazena_preco
        armazena_nome = []
        armazena_preco = []
        contador = 1
        while True:
            try:
                scraping = {
                    'XP':{
                        'nome': f'/html/body/main/section[3]/div[1]/div[2]/div[2]/div[4]/div/div[{contador}]/a/h2',
                        'preco': f'/html/body/main/section[3]/div[1]/div[2]/div[2]/div[4]/div/div[{contador}]/a/div[3]/span[1]',
                    }
                }
                
                nome = self.driver.find_element(By.XPATH, scraping['XP']['nome'])
                preco = self.driver.find_element(By.XPATH, scraping['XP']['preco'])
                nome_text = nome.text
                preco_text = preco.text

                armazena_nome.append(nome_text)
                armazena_preco.append(preco_text)

                self.driver.execute_script('arguments[0].scrollIntoView();', nome)

                print(nome_text)
                print(preco_text)

                contador +=1
                print(contador)
                sleep(0.5)   

            except NoSuchElementException:
                print('nao tem mais elementos na pagina')
                break

            except Exception as e:
                print(f'error {e}')
        
    def cria_planilhas(self):
        planilha = openpyxl.Workbook()
        atacado = planilha.active
        atacado.title = 'mercearia'
        atacado['A1'] = 'Nome'
        atacado['B1'] = 'Preco'

        for index, (nome, preco) in enumerate(zip(armazena_nome, armazena_preco), start=2):
            atacado.cell(column=1, row=index, value=nome)
            atacado.cell(column=2, row=index, value=preco)

        planilha.save('planilha_de_precos_atacadao.xlsx')
        print('planilha salva com sucesso!')

atacadao = Atacadao()
atacadao.main()
